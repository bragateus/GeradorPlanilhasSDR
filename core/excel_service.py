"""Lógica de leitura da base e geração da planilha de saída.

Regras importantes de design:
- A base original NUNCA é sobrescrita/alterada. Só é lida (pandas.read_excel).
  Isso elimina o principal risco do programa antigo: perder a base inteira
  numa falha de gravação.
- Cada geração cria um arquivo novo, com nome único (timestamp com segundos),
  então nunca sobrescreve silenciosamente uma planilha já gerada.
- A formatação (texto do "Documento", dropdown de Status) é aplicada no MESMO
  workbook antes de salvar uma única vez — evita o dobro de I/O do script
  original (que salvava, reabria e salvava de novo).
"""
from __future__ import annotations

import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation

from author import GITHUB, INICIAIS, LINKEDIN, STAMP, VERSAO

# Colunas candidatas a "chave única" de um contato/empresa, em ordem de preferência.
CANDIDATOS_CHAVE = ["Documento", "CNPJ", "CPF", "CNPJ/CPF"]


def carregar_base(caminho: str) -> pd.DataFrame:
    """Lê a planilha base em memória. Nunca escreve nada de volta nela."""
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
    df = pd.read_excel(caminho)
    return df


def detectar_coluna_chave(df: pd.DataFrame, coluna_forcada: str = "") -> str | None:
    """Descobre qual coluna identifica unicamente cada contato/empresa.

    Se `coluna_forcada` foi definida na configuração, ela tem prioridade
    (desde que exista na base). Caso contrário, tenta os candidatos usuais.
    Retorna None se nenhuma coluna adequada for encontrada (nesse caso o
    controle de duplicados cai para um hash da linha inteira).
    """
    if coluna_forcada and coluna_forcada in df.columns:
        return coluna_forcada
    for candidato in CANDIDATOS_CHAVE:
        if candidato in df.columns:
            return candidato
    return None


def calcular_chaves(df: pd.DataFrame, coluna_chave: str | None) -> pd.Series:
    """Gera uma série de strings que identificam unicamente cada linha.

    Se houver uma coluna chave (ex: Documento), usa o valor normalizado dela.
    Caso não exista nenhuma coluna adequada, usa um hash estável do conteúdo
    da linha inteira como fallback (menos ideal, mas evita repetição mesmo
    sem uma coluna de documento).
    """
    if coluna_chave:
        return df[coluna_chave].astype(str).str.strip().str.upper()
    # fallback: hash determinístico da linha inteira
    return df.astype(str).agg("|".join, axis=1).apply(
        lambda s: str(abs(hash(s)))
    )


def formatar_documento(valor) -> str:
    """Formata CPF (11 dígitos) ou CNPJ (14 dígitos) com zeros à esquerda.

    Diferente do script original (que sempre assumia 14 dígitos/CNPJ),
    aqui detectamos o tamanho real do número para não distorcer CPFs.
    """
    if pd.isna(valor):
        return ""
    digitos = re.sub(r"\D", "", str(valor))
    if not digitos:
        return str(valor)
    if len(digitos) <= 11:
        return digitos.zfill(11)
    return digitos.zfill(14)


def sanitizar_nome_arquivo(nome: str) -> str:
    """Remove caracteres inválidos para nomes de arquivo no Windows/Linux/Mac."""
    nome = unicodedata.normalize("NFKC", nome)
    nome = re.sub(r'[\\/:*?"<>|]', "", nome)
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome


def gerar_nome_arquivo(pasta_destino: str, tipo_base: str, uf: str, usuario: str) -> str:
    """Gera um nome de arquivo único (data + hora + segundos) para nunca colidir
    com um arquivo já gerado antes, mesmo que rodado várias vezes no mesmo dia."""
    agora = datetime.now()
    data_str = agora.strftime("%y.%m.%d")
    hora_str = agora.strftime("%H%M%S")
    base_nome = sanitizar_nome_arquivo(f"{data_str} {hora_str} {tipo_base} - {uf} - {usuario}.xlsx")
    caminho = os.path.join(pasta_destino, base_nome)
    # segurança extra: se por algum motivo já existir, adiciona sufixo incremental
    contador = 1
    caminho_final = caminho
    while os.path.exists(caminho_final):
        nome_sem_ext, ext = os.path.splitext(caminho)
        caminho_final = f"{nome_sem_ext} ({contador}){ext}"
        contador += 1
    return caminho_final


def gerar_planilha_saida(
    df_sdr: pd.DataFrame,
    caminho_saida: str,
    lista_status: list[str],
    coluna_chave: str | None,
    usuario: str,
    uf: str,
    tipo_base: str,
    nome_base: str,
) -> None:
    """Escreve a planilha final (aba Contatos + Resumo) em UMA única passada,
    já com formatação de texto no Documento e dropdown de Status aplicados
    antes de salvar (evita reabrir/salvar o arquivo de novo)."""

    df_saida = df_sdr.copy()
    if coluna_chave and coluna_chave in df_saida.columns:
        df_saida[coluna_chave] = df_saida[coluna_chave].apply(formatar_documento)

    if "Status" not in df_saida.columns:
        df_saida["Status"] = ""

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_saida.to_excel(writer, sheet_name="Contatos", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Contatos"]

        # Formata a coluna chave (Documento/CNPJ/CPF) como texto, coluna inteira
        if coluna_chave and coluna_chave in df_saida.columns:
            col_idx = df_saida.columns.get_loc(coluna_chave) + 1
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].number_format = "@"
            for row in worksheet.iter_rows(
                min_row=2, max_row=len(df_saida) + 1, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    cell.number_format = "@"

        # Dropdown de status
        status_col_idx = df_saida.columns.get_loc("Status") + 1
        status_letter = worksheet.cell(row=1, column=status_col_idx).column_letter
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(lista_status)}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.ranges.add(f"{status_letter}2:{status_letter}{len(df_saida) + 1}")
        worksheet.add_data_validation(dv)

        # Aba de resumo
        resumo_data = {
            "Informação": ["Usuário", "UF", "Data de Criação", "Total de Contatos", "Planilha Base"],
            "Valor": [usuario, uf, datetime.now().strftime("%d/%m/%Y %H:%M"), len(df_saida), nome_base],
        }
        pd.DataFrame(resumo_data).to_excel(writer, sheet_name="Resumo", index=False)

        # Metadados discretos — visíveis só em Propriedades do arquivo no Excel
        props = workbook.properties
        props.creator = f"Gerador SDR v{VERSAO}"
        props.lastModifiedBy = INICIAIS
        props.keywords = f"{STAMP};{GITHUB};{LINKEDIN}"
        props.description = STAMP
